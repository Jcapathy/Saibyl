# PUBLIC INTERFACE
# ─────────────────────────────────────────────────────────
# process_spreadsheet(file_bytes: bytes, filename: str) -> ProcessedAsset
# ─────────────────────────────────────────────────────────
from __future__ import annotations

import csv
import io
import tempfile

import structlog

logger = structlog.get_logger()


async def process_spreadsheet(file_bytes: bytes, filename: str) -> dict:
    """Extract tabular data as structured text from CSV or XLSX."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        text = _process_csv(file_bytes)
    elif ext in ("xlsx", "xls"):
        text = _process_xlsx(file_bytes)
    else:
        text = "Unsupported spreadsheet format"

    logger.info("spreadsheet_processed", filename=filename, chars=len(text))
    return {
        "extracted_text": text,
        "metadata": {
            "file_extension": ext,
            "file_size": len(file_bytes),
        },
    }


def _process_csv(file_bytes: bytes) -> str:
    """Convert CSV to structured text."""
    text_content = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text_content))
    rows = list(reader)

    if not rows:
        return "Empty spreadsheet"

    headers = rows[0]
    lines = [f"## Spreadsheet Data ({len(rows) - 1} rows)\n"]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join("---" for _ in headers) + " |")

    for row in rows[1:101]:  # max 100 rows
        padded = row + [""] * (len(headers) - len(row))
        lines.append("| " + " | ".join(padded[:len(headers)]) + " |")

    if len(rows) > 101:
        lines.append(f"\n*... and {len(rows) - 101} more rows*")

    return "\n".join(lines)


def _process_xlsx(file_bytes: bytes) -> str:
    """Convert XLSX to structured text using openpyxl."""
    try:
        from openpyxl import load_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp.flush()
            wb = load_workbook(tmp.name, read_only=True, data_only=True)

        parts = []
        for sheet_name in wb.sheetnames[:5]:  # max 5 sheets
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(max_row=101, values_only=True))
            if not rows:
                continue

            headers = [str(c or "") for c in rows[0]]
            lines = [f"## Sheet: {sheet_name} ({ws.max_row} rows)\n"]
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("| " + " | ".join("---" for _ in headers) + " |")

            for row in rows[1:]:
                cells = [str(c or "") for c in row[:len(headers)]]
                lines.append("| " + " | ".join(cells) + " |")

            parts.append("\n".join(lines))

        wb.close()
        return "\n\n".join(parts) if parts else "Empty spreadsheet"

    except ImportError:
        logger.warning("openpyxl_not_available")
        return "XLSX processing requires openpyxl"
